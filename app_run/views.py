from django.conf import settings
from django.contrib.auth.models import User
from django.db.models import Avg, Count, Max, Min, Q, Sum
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from geopy.distance import distance
from openpyxl import load_workbook
from rest_framework import mixins, viewsets
from rest_framework.decorators import api_view
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView

from app_run.models import (AthleteInfo, CoachRating, Challenge, CollectibleItem, Position,
                            Run, Subscribe)
from app_run.serializers import (AthleteDetailSerializer,
                                 AthleteInfoSerializer, ChallengeSerializer,
                                 CoachDetailSerializer, CoachRatingSerilizer,
                                 CollectibleItemSerializer, PositionSerializer,
                                 RunSerializer, UserSerializer)


class ProgressRunItemPagination(PageNumberPagination):
    page_size_query_param = 'size'


@api_view(['POST'])
def upload_file(request):
    uploaded_file = request.FILES.get('file')
    if uploaded_file:
        wb = load_workbook(uploaded_file)
        ws = wb.active
        errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, uid, value, latitude, longitude, url = row
            serializer = CollectibleItemSerializer(data={
                'name': name,
                'uid': uid,
                'value': value,
                'latitude': latitude,
                'longitude': longitude,
                'picture': url
            })
            if serializer.is_valid():
                CollectibleItem.objects.create(**serializer.validated_data)
            else:
                errors.append(row)
        return Response(errors)
    return Response("File is not available!", status=400)


@api_view(['GET'])
def company_details_view(request):
    details = {
        'company_name': settings.COMPANY_NAME,
        'slogan': settings.SLOGAN,
        'contacts': settings.CONTACTS
    }
    return Response(details)


class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.select_related('athlete').all()
    serializer_class = RunSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'athlete']
    ordering_fields = ['created_at']
    pagination_class = ProgressRunItemPagination


class UsersViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.annotate(
        runs_finished=Count('runs', filter=Q(runs__status='finished')),
        rating=Avg('ratings_to_coach__rating')
    )
    serializer_class = UserSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['first_name', 'last_name']
    ordering_fields = ['date_joined']
    pagination_class = ProgressRunItemPagination

    def get_queryset(self):
        type = self.request.query_params.get('type', '')
        qs = self.queryset.filter(is_superuser=False)
        if type == 'coach':
            return qs.filter(is_staff=True)
        elif type == 'athlete':
            return qs.filter(is_staff=False)
        return qs

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return CoachDetailSerializer if self.get_object().is_staff else AthleteDetailSerializer
        return super().get_serializer_class()


class AthleteInfoViewSet(mixins.RetrieveModelMixin,
                         mixins.UpdateModelMixin,
                         viewsets.GenericViewSet):
    serializer_class = AthleteInfoSerializer

    def get_object(self):
        user = get_object_or_404(User, id=self.kwargs.get('pk'))
        athlete, created = AthleteInfo.objects.get_or_create(user=user)
        return athlete

    def perform_update(self, serializer):
        user = get_object_or_404(User, id=self.kwargs.get('pk'))
        goals = serializer.validated_data.get('goals')
        weight = serializer.validated_data.get('weight')
        athlete, created = AthleteInfo.objects.update_or_create(
            user=user,
            defaults={
            'goals': goals,
            'weight': weight
        })
        return athlete

    def update(self, request, *args, **kwargs):
        super().update(request, *args, **kwargs)
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=201)


class ChallengeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['athlete']


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['run']

    def perform_create(self, serializer):
        run = serializer.validated_data.get('run')
        athlete_latitude = serializer.validated_data.get('latitude')
        athlete_longitude = serializer.validated_data.get('longitude')
        items = CollectibleItem.objects.all()
        for item in items:
            distance_to_item = round(
                distance((item.latitude, item.longitude), (athlete_latitude, athlete_longitude)).m,
                3
            )
            if distance_to_item <= 100 and not item in run.athlete.items.all():
                run.athlete.items.add(item)
        previous_position = run.positions.all().order_by('-date_time').first()
        if previous_position:
            distance_to_previous_position = round(distance((previous_position.latitude, previous_position.longitude),
                                                           (athlete_latitude, athlete_longitude)).m, 2)
            time_from_previous_position = (
                serializer.validated_data.get('date_time') - previous_position.date_time
            ).total_seconds()
            serializer.validated_data['speed'] = round(distance_to_previous_position / time_from_previous_position, 2)
            serializer.validated_data['distance'] = previous_position.distance + (distance_to_previous_position / 1000)
        serializer.save()


class CollectibleItemViewSet(viewsets.ModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer


class RunStartView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status != 'init':
            return Response({'message': 'Incorrect Status'}, status=400)
        run.status = 'in_progress'
        run.save()
        return Response(RunSerializer(run).data)


class RunStopView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status != 'in_progress':
            return Response({'message': 'Incorrect Status'}, status=400)
        run.distance = round(distance(*[(position.latitude, position.longitude)
                                        for position in run.positions.all()])
                                        .km, 3)
        timestampts = run.positions.aggregate(start=Min('date_time'), stop=Max('date_time'))
        start, stop = timestampts.get('start'), timestampts.get('stop')
        run.run_time_seconds = (stop - start).total_seconds() if start and stop else 0
        speed = run.positions.aggregate(Avg('speed')).get('speed__avg')
        run.speed = speed if speed else 0
        run.status = 'finished'
        run.save()
        athlete = run.athlete
        finished_runs = athlete.runs.filter(status='finished')
        if (finished_runs.count() >= 10
            and not Challenge.objects.filter(athlete=athlete, full_name='Сделай 10 Забегов!').exists()):
            Challenge.objects.create(full_name='Сделай 10 Забегов!', athlete=athlete)
        total_distance = finished_runs.aggregate(Sum('distance'))
        if (total_distance.get('distance__sum') >= 50
            and not Challenge.objects.filter(athlete=athlete, full_name='Пробеги 50 километров!').exists()):
            Challenge.objects.create(full_name='Пробеги 50 километров!', athlete=athlete)
        if (run.distance >= 2 and run.run_time_seconds <= 600
            and not Challenge.objects.filter(athlete=athlete, full_name='2 километра за 10 минут!').exists()):
            Challenge.objects.create(full_name='2 километра за 10 минут!', athlete=athlete)
        return Response(RunSerializer(run).data)


class SubscribeToCoachView(APIView):
    def post(self, request, id):
        print(request.data)
        athlete_id = request.data.get('athlete')
        if not athlete_id:
            return Response({ 'message': 'There isn\'t Athlete ID in request.' }, status=400)
        try:
            athlete = User.objects.get(id=athlete_id, is_superuser=False)
        except User.DoesNotExist:
            return Response({ 'message': 'Athlete instance does not exist.' }, status=400)
        if athlete.is_staff:
            return Response({ 'message': 'Only athletes could subscribe.' }, status=400)
        coach = get_object_or_404(User, id=id, is_superuser=False)
        if not coach.is_staff:
            return Response({ 'message': 'Coach Instance doesn\'t exist.' }, status=400)
        if Subscribe.objects.filter(athlete=athlete, coach=coach).exists():
            return Response({ 'message': 'Subscribe is exists.' }, status=400)
        Subscribe.objects.create(athlete=athlete, coach=coach)
        return Response({ 'message': f'{athlete} successfully subcribed to {coach}.' }, status=200)


class ChallengesSummaryView(APIView):
    def get(self, request):
        challenges = Challenge.objects.all().select_related('athlete')
        challenge_names = set([challenge.full_name for challenge in challenges])
        challenge_names = [{ 'name_to_display': challenge_name, 'athletes': [] } for challenge_name in challenge_names]
        for challenge_name in challenge_names:
            for challenge in challenges:
                if challenge_name['name_to_display'] != challenge.full_name:
                    continue
                full_name = ' '.join([challenge.athlete.first_name, challenge.athlete.last_name]) if (
                    challenge.athlete.first_name and challenge.athlete.last_name
                ) else ''
                athlete = {
                    'id': challenge.athlete.id,
                    'full_name': full_name,
                    'username': challenge.athlete.username
                }
                challenge_name['athletes'].append(athlete)
        return Response(challenge_names)


class CoachRatingsView(APIView):
    def post(self, request, coach_id):
        coach = get_object_or_404(User, id=coach_id, is_superuser=False)
        if not coach.is_staff:
            return Response({ 'message': f'The user type ({coach}) is incorrect.' }, status=400)
        athlete_id = request.data.get('athlete')
        if not athlete_id:
            return Response({ 'message': 'There isn\'t Athlete ID in request.' }, status=400)
        try:
            athlete = User.objects.get(id=athlete_id, is_superuser=False)
        except User.DoesNotExist:
            return Response({ 'message': 'Athlete instance does not exist.' }, status=400)
        if athlete.is_staff:
            return Response({ 'message': f'The user type ({athlete}) is incorrect.' }, status=400)
        if not Subscribe.objects.filter(athlete=athlete, coach=coach).exists():
            return Response({ 'message': 'Athlete ins\'t subscribed to Coach.' }, status=400)
        rating = request.data.get('rating')
        if not rating:
            return Response({ 'message': 'There isn\'t rating field in request.' }, status=400)
        serializer = CoachRatingSerilizer(data={ 'athlete': athlete.id, 'coach': coach.id, 'rating': rating })
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)
        mark, created = CoachRating.objects.update_or_create(athlete=athlete, defaults={ 'coach': coach, 'rating': rating })
        if created:
            return Response({ 'message': f'{athlete} successfully set rating ({mark.rating}) to {coach}.' }, status=200)
        return Response({ 'message': f'{athlete} successfully updated rating ({mark.rating}) to {coach}.' }, status=200)


class AnalyticsForCoachView(APIView):
    def get(self, request, coach_id):
        try:
            coach = User.objects.get(id=coach_id, is_superuser=False)
        except User.DoesNotExist:
            return Response({ 'message': f'User instance with ID: {coach_id} isn\'t exists.' })
        if not coach.is_staff:
            return Response({ 'message': f'User with ID: {coach_id} has invalid type.' })

        athletes = User.objects.filter(subscribers__coach=coach).annotate(
            longest_run=Max('runs__distance', filter=Q(runs__status='finished')),
            total_run=Sum('runs', filter=Q(runs__status='finished')),
            total_distance=Sum('runs__distance', filter=Q(runs__status='finished')),
            speed_avg=Avg('runs__speed', filter=Q(runs__status='finished'))
        )

        longest_run_user = athletes.order_by('-longest_run').first()
        total_run_user = athletes.order_by('-total_run').first()
        speed_avg_user = athletes.order_by('-speed_avg').first()

        analytics = {
            'longest_run_user': longest_run_user.id,
            'longest_run_value': longest_run_user.longest_run,
            'total_run_user': total_run_user.id,
            'total_run_value': total_run_user.total_distance,
            'speed_avg_user': speed_avg_user.id,
            'speed_avg_value': speed_avg_user.speed_avg
        }

        return Response(analytics)